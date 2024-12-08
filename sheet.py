import utils
import openpyxl
import os
import time
import uuid

from utils import Strings, Colors
from openpyxl import utils, Workbook
from openpyxl.styles import Font, Alignment, Side, Border
from openpyxl.styles.fills import PatternFill
from platform import system
from subprocess import Popen
from typing import Self

INSTRUCTION = 1
HEADER = 2
ENTRY = 3


class Sheet:

    def __init__(self):
        """
        Initialize sheet object.

        Args:
            self:   The instance of this class.
        Return
            None
        """
        self.filename = str(uuid.uuid4()) + ".xlsx"
        self.filepath = os.getcwd()

    def rename(self,
               input_dir: str,
               entries: list,
               headers: list
               ) -> Self:
        """
        This opens a spreadsheet for renaming files or folders.

        Args:
            self:               The instance of this class.
            input_dir (str):    The user-defined directory of the files and folders.
            entries (list):     The list of files of folders to manage.
            headers (list):     The headers of the spreadsheet.
        Returns:
            Self:               The instance of this class.
        """
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = Strings.spreadsheet_name

        headers_len = len(headers)

        # Row 1 = add instructions
        sheet.merge_cells(f"A1:{utils.get_column_letter(headers_len)}1")
        cell = sheet["A1"]
        cell.value = Strings.instructions_rename.format(
            working_dir=input_dir,
            filename=self.filename,
            filename_dir=self.filepath)
        cell.font = Font(
            name=font,
            color=Colors.gray_dark,
            italic=True)
        cell.alignment = Alignment(
            horizontal="left",
            vertical="center")
        sheet.row_dimensions[INSTRUCTION].height = 100

        # Row 2 = add header names
        for col, header in enumerate(headers, start=1):  # Start at column 1 since there is no 0
            cell = sheet.cell(row=HEADER, column=col)
            cell.value = header
            cell.font = Font(
                name=font,
                color=Colors.white)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center")
            cell.border = border_style
            cell.fill = PatternFill(
                fgColor=Colors.blue_dodger,
                fill_type="solid")
            sheet.column_dimensions[utils.get_column_letter(col)].width = 25
        sheet.row_dimensions[HEADER].height = 26

        # Row 3 = add data
        for row, entry in enumerate(entries, start=ENTRY):
            for col, value in enumerate(entry.values(), start=1):  # Start at column 1 since there is no 0

                if col == 1:  # Name
                    font_color = Colors.gray_dark
                    horizontal_alignment = "left"
                elif col == headers_len:  # Rename
                    font_color = Colors.purple
                    horizontal_alignment = "left"
                else:  # Details
                    font_color = Colors.gray_dark_medium
                    horizontal_alignment = "right"

                cell = sheet.cell(row=row, column=col)
                cell.value = value
                cell.font = Font(
                    name=font,
                    color=font_color)
                cell.alignment = Alignment(
                    horizontal=horizontal_alignment,
                    vertical="center"
                )
                cell.border = border_style
                cell.number_format = "@"  # Format cells as a string

        # Save the spreadsheet
        workbook.save(self.filename)
        return self

    def create(self,
               input_dir: str,
               headers: list) -> Self:
        """
        This opens a spreadsheet for renaming files or folders.

        Args:
            self:               The instance of this class.
            input_dir (str):    The user-defined directory of the files and folders.
            headers (list):     The list of files of folders to manage.
        Returns:
            Self:               The instance of this class.
        """
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = Strings.spreadsheet_name

        sheet.column_dimensions["A"].width = 67

        # Row 1 = add instructions
        sheet.row_dimensions[INSTRUCTION].height = 100
        cell_instruction = sheet["A1"]
        cell_instruction.value = Strings.instructions_create.format(
            working_dir=input_dir,
            filename=self.filename,
            filename_dir=self.filepath)
        cell_instruction.font = Font(
            name=font,
            color=Colors.gray_dark,
            italic=True)
        cell_instruction.alignment = Alignment(
            horizontal="left",
            vertical="center")

        # Row 2 = add header names
        sheet.row_dimensions[HEADER].height = 25
        cell_header = sheet["A2"]
        cell_header.value = headers[0]
        cell_header.font = Font(
            name=font,
            color=Colors.white)
        cell_header.alignment = Alignment(
            horizontal="center",
            vertical="center")
        cell_header.border = border_style
        cell_header.fill = PatternFill(
            fgColor=Colors.blue_dodger,
            fill_type="solid")

        for row in range(ENTRY, 12):  # Initial empty rows
            cell = sheet.cell(row, 1)
            cell.value = ""
            cell.font = Font(
                name=font,
                color=Colors.purple)
            cell.border = border_style
            cell.number_format = "@"  # Format cells as a string

        workbook.save(self.filename)
        return self

    def show(self) -> Self:
        """
        Opens an application for this spreadsheet.

        Args:
            self:   The instance of this class.
        Returns:
            Self:   the instance of this class.
        """
        file_name = self.filename
        match system():
            case "Darwin":
                Popen(["open", file_name])  # macOS
            case "Linux":
                Popen(["libreoffice", file_name])
            case "Windows":
                Popen(["excel.exe", file_name])
            case _:
                print("Please open the spreadsheet manually.")
                return self

        print(Strings.init_sheet)
        time.sleep(3)  # Set delay for opening application
        return self

    def load(self) -> Workbook:
        """
        Loads the saved spreadsheet file.

        Args:
            self:       The instance of this class.
        Returns:
            Workbook:   The spreadsheet object.
        """
        row_start = ENTRY  # Data starts at row 3
        file_path = os.path.join(self.filepath, self.filename)
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        return workbook.active.iter_rows(values_only=True, min_row=row_start)

    def delete(self) -> None:
        """
        Deletes the saved spreadsheet file.

        Args:
            self:   The instance of this class.
        Returns:
            None
        """
        file_path = os.path.join(self.filepath, self.filename)
        try:
            os.remove(file_path)
        except FileNotFoundError:
            print(f"Folder '{file_path}' not found.")


border_style = Border(left=Side(style="thin", color=Colors.gray_light),
                      right=Side(style="thin", color=Colors.gray_light),
                      top=Side(style="thin", color=Colors.gray_light),
                      bottom=Side(style="thin", color=Colors.gray_light))

font = "Verdana"
